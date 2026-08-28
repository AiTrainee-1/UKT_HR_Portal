"""
Attendance page support views: Skipped, Punch View, Sync status, Errors.

These replace the three device-dependent features (Sync Biometric, Manual
Import, Auto Sync) for a cloud-hosted backend. The key difference: nothing
here ever contacts a biometric device. Every endpoint reads data already in
the database -put there by the ADMS push path -so none of it can hang, time
out, or fail because the device sits on an unreachable private LAN.

That also means Punch View's export/import round-trip is safe to run from
anywhere, unlike the old Manual Import which had to reach the device first
just to produce the file.
"""
from __future__ import annotations

import io
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response

from .audit_utils import log_action
from .auth import require_hr
from .branch_scope import scope_to_branch
from .device_health import device_health
from .models import Attendance, AttendanceLog, Employee, UnmatchedPunch

# Must stay in step with the import parser below -the round trip only works
# if what we write is exactly what we expect to read back.
PUNCH_EXPORT_HEADERS = [
    "Punch ID", "Employee Code", "Employee Name", "Department",
    "Date", "Punch Time", "Punch Type", "Source",
]


def _error(message: str, code: int = 400) -> Response:
    return Response({"error": message}, status=code)


# ── Skipped: device IDs with no matching employee ────────────────────────

@api_view(["GET"])
@require_hr
def skipped_punches(request: Request) -> Response:
    """Device user IDs that punched but match no Employee.

    ?includeResolved=1 to also list ones already dealt with.
    """
    qs = UnmatchedPunch.objects.all()
    if request.query_params.get("includeResolved") not in ("1", "true", "True"):
        qs = qs.filter(resolved=False)

    rows = [{
        "id": u.id,
        "deviceUserId": u.device_user_id,
        "deviceLabel": u.device_label or None,
        "deviceSerial": u.device_serial or None,
        "punchCount": u.punch_count,
        "firstSeenAt": u.first_seen_at.isoformat() if u.first_seen_at else None,
        "lastSeenAt": u.last_seen_at.isoformat() if u.last_seen_at else None,
        "lastPunchDate": str(u.last_punch_date) if u.last_punch_date else None,
        "lastPunchTime": u.last_punch_time.strftime("%H:%M:%S") if u.last_punch_time else None,
        "resolved": u.resolved,
        "resolvedNote": u.resolved_note or None,
    } for u in qs]

    return Response({
        "results": rows,
        "unresolvedCount": UnmatchedPunch.objects.filter(resolved=False).count(),
        # Total punches being discarded -the number that actually conveys
        # the cost of leaving these unresolved.
        "discardedPunches": sum(u.punch_count for u in UnmatchedPunch.objects.filter(resolved=False)),
    })


@api_view(["POST"])
@require_hr
def resolve_skipped_punch(request: Request, pk: int) -> Response:
    """Mark one unmatched ID as dealt with. Kept, not deleted -if that ID
    punches again it un-resolves itself and reappears (see
    device_health.record_unmatched_punch)."""
    row = UnmatchedPunch.objects.filter(pk=pk).first()
    if not row:
        return _error("Not found", 404)
    row.resolved = True
    row.resolved_note = str(request.data.get("note") or "").strip()[:500]
    row.save(update_fields=["resolved", "resolved_note"])
    log_action(
        request, "update", "attendance", record_id=row.id,
        description=f"Resolved unmatched device ID {row.device_user_id}",
    )
    return Response({"ok": True})


# ── Punch View: everything already in the database ───────────────────────

def _filtered_punches(request: Request):
    """Shared filter for the list and the Excel export, so a download can
    never disagree with what's on screen."""
    # Person-wise, not time-wise: all of one employee's punches stay together,
    # in the order they happened, before the next employee begins. A purely
    # chronological feed interleaves everybody, so checking whether one
    # person's day looks right means hunting their rows out of hundreds.
    #
    # Name → date → time, all ascending. Ascending time matters: within a day
    # the rows then read as the actual sequence of the shift (first IN, last
    # OUT), which is how you spot a missing punch.
    #
    # The Excel export shares this function, so a download always matches the
    # on-screen order rather than quietly re-sorting.
    qs = (
        AttendanceLog.objects
        .select_related("employee", "employee__department")
        .order_by(
            "employee__first_name", "employee__last_name", "employee_id",
            "date", "punch_time",
        )
    )
    qs = scope_to_branch(qs, request, field="employee__branch_id")

    p = request.query_params
    if p.get("dateFrom"):
        qs = qs.filter(date__gte=p["dateFrom"])
    if p.get("dateTo"):
        qs = qs.filter(date__lte=p["dateTo"])
    if p.get("employeeId"):
        qs = qs.filter(employee_id=p["employeeId"])
    if p.get("employmentType") in ("staff", "production"):
        qs = qs.filter(employee__employment_type=p["employmentType"])
    if p.get("punchType") in (AttendanceLog.PUNCH_IN, AttendanceLog.PUNCH_OUT):
        qs = qs.filter(punch_type=p["punchType"])
    if p.get("source"):
        qs = qs.filter(source__icontains=p["source"])
    if p.get("search"):
        from django.db.models import Q
        s = p["search"].strip()
        qs = qs.filter(
            Q(employee__employee_code__icontains=s)
            | Q(employee__first_name__icontains=s)
            | Q(employee__last_name__icontains=s)
        )
    return qs


def _punch_row(log: AttendanceLog) -> dict:
    emp = log.employee
    return {
        "id": log.id,
        "employeeId": emp.id,
        "employeeCode": emp.employee_code,
        "employeeName": f"{emp.first_name} {emp.last_name}".strip(),
        "department": emp.department.name if emp.department else None,
        "employmentType": emp.employment_type,
        "date": str(log.date),
        "punchTime": log.punch_time.strftime("%H:%M:%S"),
        "punchType": log.punch_type,
        "source": log.source,
    }


@api_view(["GET"])
@require_hr
def punch_list(request: Request) -> Response:
    """Paged punch list for the Punch View table. Reads only the database —
    never touches a device, so it works regardless of where the backend runs."""
    qs = _filtered_punches(request)
    total = qs.count()

    try:
        limit = min(int(request.query_params.get("limit", 200)), 1000)
        offset = int(request.query_params.get("offset", 0))
    except ValueError:
        return _error("limit/offset must be numbers")

    return Response({
        "total": total,
        "limit": limit,
        "offset": offset,
        "results": [_punch_row(l) for l in qs[offset:offset + limit]],
    })


@api_view(["GET"])
@require_hr
def punch_export(request: Request) -> HttpResponse:
    """The current filter, as .xlsx. Includes Punch ID because the re-import
    below matches on it -that's what makes editing a downloaded file and
    uploading it back an update rather than a duplicate insert."""
    qs = _filtered_punches(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Punches"
    ws.append(PUNCH_EXPORT_HEADERS)

    fill = PatternFill("solid", fgColor="006496")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    count = 0
    for log in qs.iterator(chunk_size=1000):
        emp = log.employee
        ws.append([
            log.id,
            str(emp.employee_code or ""),
            f"{emp.first_name} {emp.last_name}".strip(),
            emp.department.name if emp.department else "",
            str(log.date),
            log.punch_time.strftime("%H:%M:%S"),
            log.punch_type,
            log.source or "",
        ])
        count += 1

    # Employee Code as text so codes like "007" keep their leading zero.
    for col, width in zip("ABCDEFGH", (10, 16, 28, 20, 14, 14, 12, 24)):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["B"].number_format = "@"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"punches-{timezone.localdate().isoformat()}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_action(request, "export", "attendance", description=f"Exported {count} punches")
    return resp


@api_view(["POST"])
@require_hr
@parser_classes([MultiPartParser, FormParser])
def punch_import(request: Request) -> Response:
    """Re-import an edited export.

    Rows carrying a Punch ID update that punch; rows with the ID blank are
    treated as new punches. Nothing is deleted -removing a row from the sheet
    does not remove the punch, because a partial export (any filter narrower
    than "everything") would otherwise wipe attendance the user never intended
    to touch.
    """
    f = request.FILES.get("file")
    if not f:
        return _error("No file uploaded. Send as multipart/form-data with key 'file'.")

    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:  # noqa: BLE001
        return _error(f"Could not read that Excel file: {exc}")

    if not rows:
        return _error("The uploaded file is empty.")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    while header and header[-1] == "":
        header.pop()
    if header != PUNCH_EXPORT_HEADERS:
        return _error(
            "Unrecognised columns. Upload a file downloaded from Punch View "
            "(edit it in place -don't add, remove or reorder columns)."
        )

    updated = created = unchanged = 0
    errors: list[str] = []

    for idx, raw in enumerate(rows[1:], start=2):
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = dict(zip(PUNCH_EXPORT_HEADERS, raw))

        code = str(row.get("Employee Code") or "").strip()
        date_raw = row.get("Date")
        time_raw = row.get("Punch Time")
        ptype = str(row.get("Punch Type") or "").strip().upper()

        if ptype not in (AttendanceLog.PUNCH_IN, AttendanceLog.PUNCH_OUT):
            errors.append(f"Row {idx}: Punch Type must be IN or OUT (got {ptype!r})")
            continue

        emp = Employee.objects.filter(employee_code=code).first()
        if not emp:
            errors.append(f"Row {idx}: no employee with code {code!r}")
            continue

        try:
            pdate = date_raw.date() if isinstance(date_raw, datetime) else datetime.strptime(str(date_raw).strip()[:10], "%Y-%m-%d").date()
            if isinstance(time_raw, datetime):
                ptime = time_raw.time()
            elif hasattr(time_raw, "hour"):
                ptime = time_raw
            else:
                ptime = datetime.strptime(str(time_raw).strip()[:8], "%H:%M:%S").time()
            ptime = ptime.replace(microsecond=0)
        except (ValueError, TypeError, AttributeError):
            errors.append(f"Row {idx}: bad Date/Punch Time ({date_raw!r} {time_raw!r}) -expected YYYY-MM-DD and HH:MM:SS")
            continue

        punch_id = row.get("Punch ID")
        existing = AttendanceLog.objects.filter(pk=int(punch_id)).first() if str(punch_id or "").strip().isdigit() else None

        if existing:
            if (existing.employee_id == emp.id and existing.date == pdate
                    and existing.punch_time == ptime and existing.punch_type == ptype):
                unchanged += 1
                continue
            # Guard the unique constraint rather than letting the DB raise:
            # an edit that collides with a punch that already exists should
            # report the conflict, not 500.
            clash = AttendanceLog.objects.filter(
                employee=emp, date=pdate, punch_time=ptime, punch_type=ptype,
            ).exclude(pk=existing.pk).exists()
            if clash:
                errors.append(f"Row {idx}: that punch already exists for {code} -edit skipped")
                continue
            existing.employee = emp
            existing.date = pdate
            existing.punch_time = ptime
            existing.punch_type = ptype
            existing.save(update_fields=["employee", "date", "punch_time", "punch_type"])
            updated += 1
        else:
            _log, was_created = AttendanceLog.objects.get_or_create(
                employee=emp, date=pdate, punch_time=ptime, punch_type=ptype,
                defaults={"source": "manual:punch-view"},
            )
            if was_created:
                created += 1
            else:
                unchanged += 1

        Attendance.objects.update_or_create(
            employee=emp, date=str(pdate), defaults={"present": True},
        )

    log_action(
        request, "update", "attendance",
        description=f"Punch View import: {updated} updated, {created} created, {len(errors)} rejected",
    )
    return Response({
        "ok": True,
        "updated": updated,
        "created": created,
        "unchanged": unchanged,
        "errors": errors[:100],
        "errorCount": len(errors),
    })


# ── Sync status + Errors ─────────────────────────────────────────────────

@api_view(["GET"])
@require_hr
def sync_status(request: Request) -> Response:
    """Live device health, for the blinking Sync indicator and Errors view.

    Cheap enough to poll: a handful of rows plus recent-punch counts, no
    device contact.
    """
    health = device_health()

    today = timezone.localdate()
    todays_punches = AttendanceLog.objects.filter(date=today).count()
    latest = AttendanceLog.objects.order_by("-date", "-punch_time").first()

    health.update({
        "punchesToday": todays_punches,
        "lastPunchAt": (
            f"{latest.date} {latest.punch_time.strftime('%H:%M:%S')}" if latest else None
        ),
        "unresolvedSkipped": UnmatchedPunch.objects.filter(resolved=False).count(),
    })
    return Response(health)
