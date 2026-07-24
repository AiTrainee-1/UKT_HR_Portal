import calendar
import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal


import bcrypt
from django.conf import settings
from django.db.models import Count, DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, throttle_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle

from .auth import require_auth, require_hr, get_token_employee_id
from .branch_scope import get_branch_scope, scope_to_branch
from .jwt_utils import sign_token
from .audit_utils import log_action, _get_ip
from .models import (
    Applicant,
    Attendance,
    AttendanceLog,
    Department,
    Employee,
    EmployeePermission,
    HRUser,
    HrLoginAttempt,
    Job,
    LeaveBalance,
    LeaveRequest,
    Notification,
    OnDutyPunchVerification,
    OnDutySession,
    Payroll,
    PushToken,
    SalaryRecord,
    SessionConfig,
    WorkSession,
)
from .serializers import (
    applicant_json,
    attendance_json,
    department_json,
    employee_full_name,
    employee_json,
    job_json,
    leave_request_json,
    notification_json,
    parse_decimal,
    salary_record_json,
)
from .payroll_views import (
    session_configs,
    session_config_detail,
    attendance_logs,
    process_punch_sessions,
    work_sessions,
    work_session_detail,
    payroll_list,
    generate_payroll,
    payroll_detail,
)



def _error(message: str, code: int = 400) -> Response:
    return Response({"error": message}, status=code)



def _employee_name(emp_id: int) -> str | None:
    emp = Employee.objects.filter(id=emp_id).first()
    return employee_full_name(emp) if emp else None


# --- Health ---


@api_view(["GET"])
def healthz(_request: Request) -> Response:
    return Response({"status": "ok"})


# --- Auth ---

# HR Portal login lockout — independent of the DRF per-IP throttle below.
# This is per-username, so an attacker rotating IPs still gets locked out.
HR_LOCKOUT_THRESHOLD = 5
HR_LOCKOUT_WINDOW_MINUTES = 15
HR_LOCKOUT_DURATION_MINUTES = 15


def _hr_username_locked_out(username: str) -> bool:
    if not username:
        return False
    window_start = timezone.now() - timedelta(minutes=HR_LOCKOUT_WINDOW_MINUTES)
    recent = HrLoginAttempt.objects.filter(
        username__iexact=username, created_at__gte=window_start
    ).order_by("-created_at")[:HR_LOCKOUT_THRESHOLD]
    if len(recent) < HR_LOCKOUT_THRESHOLD:
        return False
    # Locked out only if the most recent N attempts were ALL failures —
    # a single success resets the count.
    return all(not a.success for a in recent)


@api_view(["POST"])
@throttle_classes([ScopedRateThrottle])
def hr_login(request: Request) -> Response:
    request.throttle_scope = "login"
    username = (request.data.get("username") or "").strip()
    password = request.data.get("password") or ""

    if _hr_username_locked_out(username):
        log_action(request, "login_blocked", "auth", description=f"Locked-out login attempt for: {username}")
        return _error(
            f"Too many failed attempts. Try again in {HR_LOCKOUT_DURATION_MINUTES} minutes.", 429
        )

    account = HRUser.objects.filter(username__iexact=username, is_active=True).first()
    valid = bool(account) and bool(password) and bcrypt.checkpw(password.encode(), account.password_hash.encode())

    HrLoginAttempt.objects.create(username=username, ip_address=_get_ip(request), success=valid)

    if not valid:
        log_action(request, "login_failed", "auth", description=f"Failed login for: {username}")
        return _error("Invalid credentials", 401)

    label = account.full_name or account.username
    account.last_login = timezone.now()
    account.save(update_fields=["last_login"])
    # Shorter-lived token than employee sessions — this is the privileged portal.
    # Permissions are NOT baked into the token — see permission_middleware.py,
    # which re-checks HRUser.is_active/role.permissions fresh on every request
    # so an Admin revoking access takes effect immediately, not after expiry.
    token_payload = {
        "role": "hr",
        "name": label,
        "username": account.username,
        "hrUserId": account.id,
        "isSuperAdmin": account.is_super_admin,
    }
    token = sign_token(token_payload, expires_in=timedelta(hours=12))
    request.jwt_user = token_payload
    log_action(request, "login", "auth", description=f"{label} ({account.username}) logged in")
    return Response({"token": token, "role": "hr", "employeeId": None, "name": label})


@api_view(["POST"])
def employee_login(request: Request) -> Response:
    identifier = request.data.get("identifier")
    password = request.data.get("password")
    employee = Employee.objects.filter(
        Q(phone=identifier) | Q(email=identifier) | Q(employee_code=identifier)
    ).first()
    if not employee:
        return _error("You are not registered. Please contact HR.", 401)
    if not employee.password_hash:
        return _error("No password set. Please set your password first.", 401)
    if not bcrypt.checkpw(password.encode(), employee.password_hash.encode()):
        return _error("Invalid password", 401)
    name = employee_full_name(employee)
    token = sign_token({"role": "employee", "employeeId": employee.id, "name": name})
    return Response(
        {"token": token, "role": "employee", "employeeId": employee.id, "name": name}
    )


@api_view(["POST"])
def set_password(request: Request) -> Response:
    identifier = request.data.get("identifier")
    password = request.data.get("password")
    if not password or len(password) < 8:
        return _error("Validation error")
    employee = Employee.objects.filter(
        Q(phone=identifier) | Q(email=identifier) | Q(employee_code=identifier)
    ).first()
    if not employee:
        return _error("Employee not found. Please contact HR.", 404)
    employee.password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt(rounds=10)).decode()
    employee.save(update_fields=["password_hash", "updated_at"])
    return Response({"message": "Password set successfully"})


@api_view(["GET"])
@require_auth
def auth_me(request: Request) -> Response:
    user = request.jwt_user
    payload = {
        "role": user.get("role"),
        "employeeId": user.get("employeeId"),
        "name": user.get("name", ""),
    }
    if user.get("role") == "hr":
        # Resolved fresh from the DB (not trusted from the token) so a
        # permission change by an Admin is reflected on the next page load.
        hr_user = (
            HRUser.objects.select_related("role", "branch")
            .filter(id=user.get("hrUserId"), is_active=True)
            .first()
        )
        is_super_admin = bool(hr_user and hr_user.is_super_admin)
        permissions = (hr_user.role.permissions if hr_user and hr_user.role else {}) or {}
        payload["isSuperAdmin"] = is_super_admin
        payload["permissions"] = permissions
        payload["branchId"] = hr_user.branch_id if hr_user else None
        payload["branchName"] = hr_user.branch.name if hr_user and hr_user.branch else None
    return Response(payload)


# --- Departments ---


@api_view(["GET", "POST"])
def departments(request: Request) -> Response:
    if request.method == "GET":
        rows = (
            Department.objects.annotate(
                employee_count=Count("employees", filter=Q(employees__status="active"))
            )
            .order_by("id")
            .values("id", "name", "description", "employee_count")
        )
        return Response(
            [
                department_json(
                    Department(id=r["id"], name=r["name"], description=r["description"]),
                    r["employee_count"],
                )
                for r in rows
            ]
        )
    wrapped = require_hr(_departments_create)
    return wrapped(request)


def _departments_create(request: Request) -> Response:
    dept = Department.objects.create(
        name=request.data.get("name"),
        description=request.data.get("description"),
    )
    return Response(department_json(dept, 0), status=201)


@api_view(["GET", "DELETE"])
@require_hr
def delete_department(request: Request, pk: int) -> Response:
    try:
        dept = Department.objects.get(pk=pk)
    except Department.DoesNotExist:
        return _error("Department not found", 404)

    if request.method == "GET":
        emp_count = dept.employees.filter(status="active").count()
        return Response(department_json(dept, emp_count))

    dept.delete()
    return Response({"message": "Department deleted"})


# --- Employees ---


def _employee_queryset():
    return Employee.objects.select_related("department", "designation", "branch")


def _serialize_employee(emp: Employee) -> dict:
    dept_name = emp.department.name if emp.department_id and emp.department else None
    return employee_json(emp, dept_name)


@api_view(["GET", "POST"])
def employees(request: Request) -> Response:
    if request.method == "GET":
        return require_auth(_employees_list)(request)
    return require_hr(_employees_create)(request)


def _employees_list(request: Request) -> Response:
    qs = _employee_queryset()
    qs = scope_to_branch(qs, request)
    dept_id = request.query_params.get("departmentId")
    desig_id = request.query_params.get("designationId")
    branch_id = request.query_params.get("branchId")
    emp_status = request.query_params.get("status")
    salary_type = request.query_params.get("salaryType")
    search = request.query_params.get("search", "").strip()
    if dept_id:
        qs = qs.filter(department_id=int(dept_id))
    if desig_id:
        qs = qs.filter(designation_id=int(desig_id))
    if branch_id:
        qs = qs.filter(branch_id=int(branch_id))
    if emp_status:
        qs = qs.filter(status=emp_status)
    if salary_type:
        qs = qs.filter(salary_type=salary_type)
    if search:
        qs = qs.filter(Q(employee_code__icontains=search) | Q(phone__icontains=search))
    return Response([_serialize_employee(e) for e in qs])


def _assign_unit_code(branch_id: int | None) -> str | None:
    """
    Next "<branch code>-<n>" identifier for a branch, e.g. HO-1, HO-2,
    Unit1-1 — atomically incremented (select_for_update, inside a
    transaction) so two concurrent requests can never be handed the same
    number, and a number is never reused even after the employee holding it
    is later deleted or moved to a different branch. None if the employee
    has no branch, or that branch has no code set yet.
    """
    if branch_id is None:
        return None

    from django.db import transaction
    from .models import Branch

    with transaction.atomic():
        branch = Branch.objects.select_for_update().filter(pk=branch_id).first()
        if branch is None or not branch.code:
            return None
        branch.next_employee_seq += 1
        branch.save(update_fields=["next_employee_seq"])
        return f"{branch.code}-{branch.next_employee_seq}"


def _resolve_employee_relations(data: dict, request: Request) -> tuple[dict, list[str]]:
    """
    Resolve department/designation/branch from either an <field>Id (int FK —
    what the Add Employee dropdowns send) or a plain <field> name string
    (what the bulk-upload Excel importer sends). Returns (kwargs, warnings)
    where kwargs has department/designation/branch_id ready for
    Employee.objects.create(), and warnings are non-fatal notes about names
    that couldn't be matched (the row/request still succeeds).
    """
    from .models import Branch, Designation as _Desig

    warnings: list[str] = []

    dept = None
    if data.get("departmentId"):
        dept = Department.objects.filter(pk=int(data["departmentId"])).first()
    elif str(data.get("department") or "").strip():
        dept_name = str(data["department"]).strip()
        dept = Department.objects.filter(name__iexact=dept_name).first()
        if dept is None:
            dept, _created = Department.objects.get_or_create(name=dept_name)

    desig = None
    if data.get("designationId"):
        desig = _Desig.objects.filter(pk=int(data["designationId"])).first()
    elif str(data.get("designation") or "").strip():
        desig_title = str(data["designation"]).strip()
        desig_qs = _Desig.objects.filter(title__iexact=desig_title)
        desig = (desig_qs.filter(department=dept).first() if dept else None) or desig_qs.first()
        if desig is None:
            warnings.append(f"Designation '{desig_title}' not found — left blank")

    # A branch-scoped HR user can only ever create employees in their own
    # branch, regardless of what the client/row sends. Unscoped users (super
    # admin, MD/Directors, branch-less HR) pick one explicitly, by id or name.
    scoped_branch_id = get_branch_scope(request)
    if scoped_branch_id is not None:
        branch_id = scoped_branch_id
    elif data.get("branchId"):
        branch_id = int(data["branchId"])
    elif str(data.get("branch") or "").strip():
        branch_name = str(data["branch"]).strip()
        b = Branch.objects.filter(name__iexact=branch_name).first()
        branch_id = b.id if b else None
        if b is None:
            warnings.append(f"Branch '{branch_name}' not found — left unassigned")
    else:
        branch_id = None

    return {"department": dept, "designation": desig, "branch_id": branch_id}, warnings


def _create_employee_from_data(
    data: dict, request: Request, strict: bool = True,
) -> tuple[Employee | None, str | None, list[str]]:
    """
    Create one Employee from a plain camelCase dict — the shape shared by
    both the single Add Employee JSON body and one row of a bulk-upload
    Excel import. Returns (employee, error, warnings): error is a hard-fail
    reason (nothing created); warnings are non-fatal notes about fields that
    were skipped (e.g. an unmatched department/branch name).

    `strict` gates Last Name / Phone as required — on for the single Add
    Employee form, off for bulk upload (where only Employee Code and First
    Name are mandatory; Last Name and Phone may be filled in later).
    """
    employee_code = str(data.get("employeeCode") or "").strip()
    first_name = str(data.get("firstName") or "").strip()
    last_name = str(data.get("lastName") or "").strip()
    phone = str(data.get("phone") or "").strip()

    if not employee_code:
        return None, "Employee code is required", []
    if not first_name:
        return None, "First name is required", []
    if strict and not last_name:
        return None, "Last name is required", []
    if strict and not phone:
        return None, "Phone is required", []
    if Employee.objects.filter(employee_code=employee_code).exists():
        return None, f"Employee code '{employee_code}' already exists", []

    relations, warnings = _resolve_employee_relations(data, request)
    unit_code = _assign_unit_code(relations["branch_id"])

    emp = Employee.objects.create(
        employee_code=employee_code,
        first_name=first_name,
        last_name=last_name,
        gender=data.get("gender") or None,
        date_of_birth=data.get("dateOfBirth") or None,
        email=data.get("email") or None,
        phone=phone,
        role=data.get("role") or None,
        employment_type=data.get("employmentType") or "staff",
        department=relations["department"],
        designation=relations["designation"],
        branch_id=relations["branch_id"],
        unit_code=unit_code,
        salary_type=data.get("salaryType") or "monthly",
        salary_amount=parse_decimal(data.get("salaryAmount")),
        salary_per_shift=parse_decimal(data.get("salaryPerShift")),
        bank_name=data.get("bankName") or None,
        bank_account=data.get("bankAccount") or None,
        bank_ifsc=data.get("bankIfsc") or None,
        pf_number=data.get("pfNumber") or None,
        esi_number=data.get("esiNumber") or None,
        id_proof=data.get("idProof") or None,
        address=data.get("address") or None,
        join_date=data.get("joinDate") or None,
        father_name=data.get("fatherName") or None,
        mother_name=data.get("motherName") or None,
        biometric_device_id=data.get("biometricDeviceId") or None,
        photo_url=data.get("photoUrl") or None,
        blood_group=data.get("bloodGroup") or None,
        emergency_contact=data.get("emergencyContact") or None,
    )
    from .shift_views import auto_assign_production_shift
    auto_assign_production_shift(emp)
    return emp, None, warnings


def _employees_create(request: Request) -> Response:
    emp, error, _warnings = _create_employee_from_data(request.data, request)
    if error:
        return _error(error)

    emp = _employee_queryset().get(pk=emp.pk)
    log_action(request, "create", "employees", record_id=emp.id,
               description=f"Created employee {emp.employee_code} — {emp.first_name} {emp.last_name}")
    return Response(_serialize_employee(emp), status=201)


@api_view(["GET", "PATCH", "DELETE"])
def employee_detail(request: Request, pk: int) -> Response:
    if request.method == "GET":
        return require_auth(_employee_get)(request, pk)
    if request.method == "PATCH":
        return require_hr(_employee_update)(request, pk)
    return require_hr(_employee_delete)(request, pk)


def _employee_get(request: Request, pk: int) -> Response:
    emp = scope_to_branch(_employee_queryset(), request).filter(pk=pk).first()
    if not emp:
        return _error("Employee not found", 404)
    return Response(_serialize_employee(emp))


def _employee_update(request: Request, pk: int) -> Response:
    emp = scope_to_branch(_employee_queryset(), request).filter(pk=pk).first()
    if not emp:
        return _error("Employee not found", 404)

    original_branch_id = emp.branch_id

    # Handle department: prefer departmentId (int FK), fall back to name string
    if "departmentId" in request.data:
        raw = request.data.get("departmentId")
        emp.department_id = int(raw) if raw else None
    elif "department" in request.data:
        dept_name = request.data.get("department", "").strip()
        if dept_name:
            dept, _ = Department.objects.get_or_create(name=dept_name)
            emp.department = dept

    # Handle designation
    if "designationId" in request.data:
        raw = request.data.get("designationId")
        emp.designation_id = int(raw) if raw else None

    # Handle branch: a branch-scoped HR user can't move an employee to
    # another branch (their own branch_id wins regardless of payload).
    scoped_branch_id = get_branch_scope(request)
    if scoped_branch_id is not None:
        emp.branch_id = scoped_branch_id
    elif "branchId" in request.data:
        raw = request.data.get("branchId")
        emp.branch_id = int(raw) if raw else None

    if emp.branch_id != original_branch_id:
        # Moved to a different branch (or removed from one) — the old Unit
        # Code no longer describes them, so retire it and mint a fresh one
        # for the new branch (never touches the old branch's counter).
        emp.unit_code = _assign_unit_code(emp.branch_id)

    if "employeeCode" in request.data:
        new_code = (request.data["employeeCode"] or "").strip()
        if not new_code:
            return _error("Employee code is required")
        if Employee.objects.filter(employee_code=new_code).exclude(pk=pk).exists():
            return _error("Employee code already exists")
        emp.employee_code = new_code

    field_map = {
        "firstName": "first_name",
        "lastName": "last_name",
        "gender": "gender",
        "dateOfBirth": "date_of_birth",
        "email": "email",
        "phone": "phone",
        "role": "role",
        "employmentType": "employment_type",
        "salaryType": "salary_type",
        "salaryAmount": "salary_amount",
        "salaryPerShift": "salary_per_shift",
        "status": "status",
        "bankName": "bank_name",
        "bankAccount": "bank_account",
        "bankIfsc": "bank_ifsc",
        "idProof": "id_proof",
        "pfNumber": "pf_number",
        "esiNumber": "esi_number",
        "address": "address",
        "joinDate": "join_date",
        "fatherName": "father_name",
        "motherName": "mother_name",
        "biometricDeviceId": "biometric_device_id",
        "photoUrl": "photo_url",
        "bloodGroup": "blood_group",
        "emergencyContact": "emergency_contact",
        "locationTrackingEnabled": "location_tracking_enabled",
    }
    if "employmentType" in request.data:
        emp_type = request.data.get("employmentType")
        if emp_type not in (Employee.EMPLOYMENT_TYPE_STAFF, Employee.EMPLOYMENT_TYPE_PRODUCTION):
            return _error("employmentType must be 'staff' or 'production'")

    for json_key, model_key in field_map.items():
        if json_key in request.data:
            value = request.data[json_key]
            if model_key in ("salary_amount", "salary_per_shift"):
                value = parse_decimal(value)
            elif model_key == "date_of_birth":
                value = value or None
            setattr(emp, model_key, value)
    emp.save()
    # If employee type was changed to production, try to auto-assign a production shift
    if request.data.get("employmentType") == "production":
        from .shift_views import auto_assign_production_shift
        emp_fresh = Employee.objects.get(pk=pk)
        auto_assign_production_shift(emp_fresh)

    emp = _employee_queryset().get(pk=pk)
    log_action(request, "update", "employees", record_id=pk,
               description=f"Updated employee {emp.employee_code} — {emp.first_name} {emp.last_name}")
    return Response(_serialize_employee(emp))


def _employee_delete(request: Request, pk: int) -> Response:
    emp = scope_to_branch(Employee.objects, request).filter(id=pk).first()
    if not emp:
        return _error("Employee not found", 404)
    name = f"{emp.employee_code} — {emp.first_name} {emp.last_name}"
    emp.delete()
    log_action(request, "delete", "employees", record_id=pk, description=f"Deleted employee {name}")
    return Response({"message": "Employee deleted"})


@api_view(["PATCH"])
@require_hr
def employee_status(request: Request, pk: int) -> Response:
    emp = scope_to_branch(_employee_queryset(), request).filter(pk=pk).first()
    if not emp:
        return _error("Employee not found", 404)
    emp.status = request.data.get("status")
    emp.save(update_fields=["status", "updated_at"])
    return Response(_serialize_employee(emp))


# --- Bulk employee upload ---
#
# Column order/text is the enforced contract with the downloaded template —
# keep this in sync with EMPLOYEE_TEMPLATE_HEADERS in
# frontend/src/pages/hr/BulkUploadEmployees.tsx if either ever changes.
EMPLOYEE_UPLOAD_HEADERS = [
    "Employee Code", "First Name", "Last Name", "Email", "Phone", "Gender",
    "Date of Birth", "Employment Type", "Department", "Designation", "Branch",
    "Salary Type", "Salary Amount", "Salary Per Shift", "Join Date",
    "Bank Name", "Bank Account", "Bank IFSC", "PF Number", "ESI Number",
    "Address", "ID Proof", "Father's Name", "Mother's Name",
    "Biometric Device ID", "Blood Group", "Emergency Contact",
]

_VALID_EMPLOYMENT_TYPES = {"staff", "production"}
_VALID_SALARY_TYPES = {"monthly", "weekly"}
_VALID_GENDERS = {"male", "female", "other"}


def _parse_date_cell(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return raw  # unparseable — let Django's own validation reject it with its own message


def _employee_row_to_data(row: dict) -> tuple[dict, str | None]:
    """Normalize one raw Excel row (header -> cell value) into the camelCase
    dict _create_employee_from_data expects. Returns (data, error) — error is
    set when a restricted-choice column (Employment Type/Salary Type/Gender)
    holds something other than one of its known values or blank."""

    def cell(key: str) -> str:
        v = row.get(key)
        return "" if v is None else str(v).strip()

    def num_cell(key: str):
        v = row.get(key)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        return v

    emp_type = cell("Employment Type").lower()
    if emp_type and emp_type not in _VALID_EMPLOYMENT_TYPES:
        return {}, f"Employment Type must be Staff or Production, got '{cell('Employment Type')}'"

    salary_type = cell("Salary Type").lower()
    if salary_type and salary_type not in _VALID_SALARY_TYPES:
        return {}, f"Salary Type must be Monthly or Weekly, got '{cell('Salary Type')}'"

    gender = cell("Gender").lower()
    if gender and gender not in _VALID_GENDERS:
        return {}, f"Gender must be Male, Female or Other, got '{cell('Gender')}'"

    return {
        "employeeCode": cell("Employee Code"),
        "firstName": cell("First Name"),
        "lastName": cell("Last Name"),
        "email": cell("Email") or None,
        "phone": cell("Phone"),
        "gender": gender or None,
        "dateOfBirth": _parse_date_cell(row.get("Date of Birth")),
        "employmentType": emp_type or None,
        "department": cell("Department") or None,
        "designation": cell("Designation") or None,
        "branch": cell("Branch") or None,
        "salaryType": salary_type or None,
        "salaryAmount": num_cell("Salary Amount"),
        "salaryPerShift": num_cell("Salary Per Shift"),
        "joinDate": _parse_date_cell(row.get("Join Date")),
        "bankName": cell("Bank Name") or None,
        "bankAccount": cell("Bank Account") or None,
        "bankIfsc": cell("Bank IFSC") or None,
        "pfNumber": cell("PF Number") or None,
        "esiNumber": cell("ESI Number") or None,
        "address": cell("Address") or None,
        "idProof": cell("ID Proof") or None,
        "fatherName": cell("Father's Name") or None,
        "motherName": cell("Mother's Name") or None,
        "biometricDeviceId": cell("Biometric Device ID") or None,
        "bloodGroup": cell("Blood Group") or None,
        "emergencyContact": cell("Emergency Contact") or None,
    }, None


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@require_hr
def bulk_upload_employees(request: Request) -> Response:
    file = request.FILES.get("file")
    if not file:
        return _error("No file uploaded. Send as multipart/form-data with key 'file'.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return _error(f"Failed to read the Excel file: {e}")

    if not rows:
        return _error("The uploaded file is empty.")

    def _normalize_header(c) -> str:
        # The downloaded template marks required columns as "Employee Code *"
        # for the user's benefit — strip that trailing marker back off before
        # comparing, so an unmodified official template always validates.
        h = str(c).strip() if c is not None else ""
        return h[:-1].rstrip() if h.endswith("*") else h

    header_row = [_normalize_header(c) for c in rows[0]]
    while header_row and header_row[-1] == "":
        header_row.pop()

    if header_row != EMPLOYEE_UPLOAD_HEADERS:
        return Response(
            {
                "error": "invalid_template",
                "message": "Invalid template. Please upload the employee data using the official template provided by the system.",
            },
            status=400,
        )

    created = 0
    sample_skipped = 0
    errors: list[str] = []
    warnings: list[str] = []

    for idx, raw_row in enumerate(rows[1:], start=2):
        if not raw_row or all(c is None or str(c).strip() == "" for c in raw_row):
            continue  # skip fully blank rows
        first_cell = str(raw_row[0]).strip() if raw_row[0] is not None else ""
        if first_cell.upper().startswith("SAMPLE"):
            # Reference rows shipped in the downloaded template — never imported,
            # even if the user forgets to delete them before uploading.
            sample_skipped += 1
            continue
        row = dict(zip(EMPLOYEE_UPLOAD_HEADERS, raw_row))
        data, row_error = _employee_row_to_data(row)
        if row_error:
            errors.append(f"Row {idx}: {row_error}")
            continue
        emp, error, row_warnings = _create_employee_from_data(data, request, strict=False)
        if error:
            errors.append(f"Row {idx}: {error}")
            continue
        created += 1
        log_action(
            request, "create", "employees", record_id=emp.id,
            description=f"Bulk-imported employee {emp.employee_code} — {emp.first_name} {emp.last_name}",
        )
        warnings.extend(f"Row {idx}: {w}" for w in row_warnings)

    return Response(
        {
            "message": f"Imported {created} employee(s)." + (f" {len(errors)} row(s) failed." if errors else ""),
            "created": created,
            "failed": len(errors),
            "sampleRowsSkipped": sample_skipped,
            "errors": errors,
            "warnings": warnings,
        },
        status=201,
    )


# --- Bulk employee update (existing employees, matched by Employee Code) ---

# Fields the Excel updater may change, as (header, model attr) pairs. Employee
# Code is deliberately absent — it's the match key, so this flow can never
# rename it. Department/Designation/Branch are handled separately (FK
# resolution), as are the choice-validated columns.
_UPDATE_TEXT_FIELDS = {
    "First Name": "first_name",
    "Last Name": "last_name",
    "Email": "email",
    "Phone": "phone",
    "Bank Name": "bank_name",
    "Bank Account": "bank_account",
    "Bank IFSC": "bank_ifsc",
    "PF Number": "pf_number",
    "ESI Number": "esi_number",
    "Address": "address",
    "ID Proof": "id_proof",
    "Father's Name": "father_name",
    "Mother's Name": "mother_name",
    "Biometric Device ID": "biometric_device_id",
    "Blood Group": "blood_group",
    "Emergency Contact": "emergency_contact",
}


def _apply_row_updates(emp, row: dict, request: Request) -> tuple[list[str], list[str], str | None]:
    """
    Apply one Excel row's non-blank cells onto an existing Employee, writing
    only values that actually differ. Blank cells always mean "leave as is" —
    this flow exists to fill gaps and fix mistakes, so an empty cell must
    never wipe stored data. Returns (changed_field_labels, warnings, error).
    Nothing is saved here; the caller saves when changes is non-empty.
    """
    from .models import Branch, Designation as _Desig

    changed: list[str] = []
    warnings: list[str] = []

    def cell(key: str) -> str:
        v = row.get(key)
        return "" if v is None else str(v).strip()

    # Choice-validated columns — a bad value fails the whole row rather than
    # silently skipping, so typos get fixed instead of ignored.
    gender = cell("Gender").lower()
    if gender and gender not in _VALID_GENDERS:
        return [], [], f"Gender must be Male, Female or Other, got '{cell('Gender')}'"
    emp_type = cell("Employment Type").lower()
    if emp_type and emp_type not in _VALID_EMPLOYMENT_TYPES:
        return [], [], f"Employment Type must be Staff or Production, got '{cell('Employment Type')}'"
    salary_type = cell("Salary Type").lower()
    if salary_type and salary_type not in _VALID_SALARY_TYPES:
        return [], [], f"Salary Type must be Monthly or Weekly, got '{cell('Salary Type')}'"

    for header, attr in _UPDATE_TEXT_FIELDS.items():
        value = cell(header)
        if value and value != str(getattr(emp, attr) or ""):
            setattr(emp, attr, value)
            changed.append(header)

    for value, attr, header in [
        (gender, "gender", "Gender"),
        (emp_type, "employment_type", "Employment Type"),
        (salary_type, "salary_type", "Salary Type"),
    ]:
        if value and value != (getattr(emp, attr) or ""):
            setattr(emp, attr, value)
            changed.append(header)

    for header, attr in [("Date of Birth", "date_of_birth"), ("Join Date", "join_date")]:
        if cell(header):
            parsed = _parse_date_cell(row.get(header))
            if parsed and parsed != str(getattr(emp, attr) or ""):
                setattr(emp, attr, parsed)
                changed.append(header)

    for header, attr in [("Salary Amount", "salary_amount"), ("Salary Per Shift", "salary_per_shift")]:
        if cell(header):
            parsed = parse_decimal(row.get(header))
            current = getattr(emp, attr)
            if parsed is not None and (current is None or Decimal(str(current)) != Decimal(str(parsed))):
                setattr(emp, attr, parsed)
                changed.append(header)

    dept_name = cell("Department")
    if dept_name and dept_name.lower() != (emp.department.name.lower() if emp.department_id and emp.department else ""):
        dept = Department.objects.filter(name__iexact=dept_name).first()
        if dept is None:
            dept, _ = Department.objects.get_or_create(name=dept_name)
        emp.department = dept
        changed.append("Department")

    desig_title = cell("Designation")
    if desig_title and desig_title.lower() != (emp.designation.title.lower() if emp.designation_id and emp.designation else ""):
        desig_qs = _Desig.objects.filter(title__iexact=desig_title)
        desig = (desig_qs.filter(department=emp.department).first() if emp.department_id else None) or desig_qs.first()
        if desig is None:
            warnings.append(f"Designation '{desig_title}' not found — kept the current one")
        else:
            emp.designation = desig
            changed.append("Designation")

    # Branch: a branch-scoped HR user can't move employees between branches
    # (same rule as the Edit Employee form) — their rows silently keep the
    # current branch. Unscoped users match by name; unknown names warn.
    branch_name = cell("Branch")
    if branch_name and get_branch_scope(request) is None:
        current_branch_name = emp.branch.name if emp.branch_id and emp.branch else ""
        if branch_name.lower() != current_branch_name.lower():
            b = Branch.objects.filter(name__iexact=branch_name).first()
            if b is None:
                warnings.append(f"Branch '{branch_name}' not found — kept the current one")
            else:
                emp.branch_id = b.id
                # Old Unit Code described the old branch — mint a fresh one.
                emp.unit_code = _assign_unit_code(b.id)
                changed.append("Branch")

    return changed, warnings, None


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@require_hr
def bulk_update_employees(request: Request) -> Response:
    """
    Companion to bulk_upload_employees for EXISTING employees: HR downloads
    the current-employees export, fills in missing/corrected cells, and
    re-uploads it here. Rows match by Employee Code; codes not in the system
    are reported (never created — that's what bulk upload is for); blank
    cells never overwrite stored data; only genuinely changed fields are
    written, and the response lists exactly what changed per employee.
    """
    file = request.FILES.get("file")
    if not file:
        return _error("No file uploaded. Send as multipart/form-data with key 'file'.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return _error(f"Failed to read the Excel file: {e}")

    if not rows:
        return _error("The uploaded file is empty.")

    def _normalize_header(c) -> str:
        h = str(c).strip() if c is not None else ""
        return h[:-1].rstrip() if h.endswith("*") else h

    header_row = [_normalize_header(c) for c in rows[0]]
    while header_row and header_row[-1] == "":
        header_row.pop()

    if header_row != EMPLOYEE_UPLOAD_HEADERS:
        return Response(
            {
                "error": "invalid_template",
                "message": "Invalid file. Please upload the Excel downloaded from Download Current Employees (or the official template) without changing its columns.",
            },
            status=400,
        )

    updated = 0
    unchanged = 0
    sample_skipped = 0
    not_found: list[str] = []
    errors: list[str] = []
    warnings: list[str] = []
    changes: list[str] = []

    scoped_qs = scope_to_branch(_employee_queryset(), request)

    for idx, raw_row in enumerate(rows[1:], start=2):
        if not raw_row or all(c is None or str(c).strip() == "" for c in raw_row):
            continue
        first_cell = str(raw_row[0]).strip() if raw_row[0] is not None else ""
        if first_cell.upper().startswith("SAMPLE"):
            sample_skipped += 1
            continue
        if not first_cell:
            errors.append(f"Row {idx}: Employee Code is required to match an existing employee")
            continue

        row = dict(zip(EMPLOYEE_UPLOAD_HEADERS, raw_row))
        emp = scoped_qs.filter(employee_code=first_cell).first()
        if emp is None:
            not_found.append(f"Row {idx}: no employee with code '{first_cell}' — use Bulk Upload to add new employees")
            continue

        changed, row_warnings, row_error = _apply_row_updates(emp, row, request)
        warnings.extend(f"Row {idx} ({first_cell}): {w}" for w in row_warnings)
        if row_error:
            errors.append(f"Row {idx} ({first_cell}): {row_error}")
            continue
        if not changed:
            unchanged += 1
            continue

        emp.save()
        if "Employment Type" in changed and emp.employment_type == "production":
            from .shift_views import auto_assign_production_shift
            auto_assign_production_shift(emp)
        updated += 1
        changes.append(f"{first_cell} — {emp.first_name} {emp.last_name}: {', '.join(changed)}")
        log_action(
            request, "update", "employees", record_id=emp.id,
            description=f"Bulk-updated employee {emp.employee_code} — changed {', '.join(changed)}",
        )

    return Response(
        {
            "message": f"Updated {updated} employee(s), {unchanged} already up to date."
            + (f" {len(not_found)} code(s) not found." if not_found else "")
            + (f" {len(errors)} row(s) failed." if errors else ""),
            "updated": updated,
            "unchanged": unchanged,
            "notFound": not_found,
            "failed": len(errors),
            "sampleRowsSkipped": sample_skipped,
            "errors": errors,
            "warnings": warnings,
            "changes": changes,
        }
    )


# --- Salary ---


def _salary_with_name(record: SalaryRecord) -> dict:
    name = _employee_name(record.employee_id)
    return salary_record_json(record, name)


def _salary_from_payroll(payroll: Payroll) -> dict:
    return {
        "id": payroll.id,
        "employeeId": payroll.employee_id,
        "month": payroll.month,
        "year": payroll.year,
        "amount": float(payroll.final_salary),
        "type": payroll.salary_mode,
        "status": payroll.status,
        "notes": payroll.notes,
        "createdAt": payroll.created_at.isoformat() if payroll.created_at else None,
    }


@api_view(["GET", "POST"])
def salary_records(request: Request) -> Response:
    if request.method == "GET":
        return require_auth(_salary_records_list)(request)
    return require_hr(_salary_records_create)(request)


def _salary_records_list(request: Request) -> Response:
    employee_id = request.query_params.get("employeeId")
    month = request.query_params.get("month")
    year = request.query_params.get("year")
    
    if not employee_id:
        return _error("employeeId required", 400)
    
    employee_id = int(employee_id)
    result = []
    
    # Fetch Payroll records (primary source)
    payroll_qs = Payroll.objects.filter(employee_id=employee_id)
    if month:
        payroll_qs = payroll_qs.filter(month=int(month))
    if year:
        payroll_qs = payroll_qs.filter(year=int(year))
    
    for p in payroll_qs:
        result.append({
            "id": p.id,
            "employeeId": p.employee_id,
            "month": p.month,
            "year": p.year,
            "amount": float(p.final_salary),
            "type": p.salary_mode,
            "status": p.status,
            "notes": p.notes,
            "createdAt": p.created_at.isoformat() if p.created_at else None,
        })
    
    # Fallback to SalaryRecord if no Payroll records exist
    if not result:
        salary_qs = SalaryRecord.objects.filter(employee_id=employee_id).select_related("employee")
        if month:
            salary_qs = salary_qs.filter(month=int(month))
        if year:
            salary_qs = salary_qs.filter(year=int(year))
        result = [_salary_with_name(r) for r in salary_qs]
    
    # Sort by year descending, then month descending
    result.sort(key=lambda x: (x["year"], x["month"]), reverse=True)
    return Response(result)


def _salary_records_create(request: Request) -> Response:
    data = request.data
    record = SalaryRecord.objects.create(
        employee_id=data.get("employeeId"),
        month=data.get("month"),
        year=data.get("year"),
        amount=parse_decimal(data.get("amount")),
        type=data.get("type", "monthly"),
        week_number=data.get("weekNumber"),
        status=data.get("status", "pending"),
        notes=data.get("notes"),
    )
    return Response(_salary_with_name(record), status=201)


@api_view(["PATCH"])
@require_hr
def update_salary_record(request: Request, pk: int) -> Response:
    record = SalaryRecord.objects.filter(pk=pk).first()
    if not record:
        return _error("Not found", 404)
    if "amount" in request.data:
        record.amount = parse_decimal(request.data["amount"])
    if "status" in request.data:
        record.status = request.data["status"]
    if "notes" in request.data:
        record.notes = request.data["notes"]
    record.save()
    return Response(_salary_with_name(record))


@api_view(["POST"])
@require_hr
def calculate_salary_records(request: Request) -> Response:
    month = request.data.get("month")
    year = request.data.get("year")
    if not month or not year:
        return _error("Month and year are required", 400)
    
    try:
        month = int(month)
        year = int(year)
    except ValueError:
        return _error("Month and year must be integers", 400)

    # Fetch all active employees
    employees = Employee.objects.filter(status="active")
    generated_count = 0
    updated_count = 0

    prefix = f"{year}-{month:02d}"

    for emp in employees:
        # Count present days for this employee in this month
        present_days = Attendance.objects.filter(
            employee=emp,
            date__startswith=prefix,
            present=True
        ).count()

        total_working_days = Attendance.objects.filter(
            employee=emp,
            date__startswith=prefix
        ).count()

        if total_working_days == 0:
            # Skip if there's no attendance record for this month
            continue

        if emp.salary_type == "monthly":
            # Per day rate based on 26 days
            per_day = emp.salary_amount / Decimal("26.00")
            calculated_amount = per_day * Decimal(present_days)
            notes = f"Auto-calculated: worked {present_days}/{total_working_days} days. Monthly Base: ₹{emp.salary_amount:,.2f}"
        else: # weekly rate
            # Per day rate based on 6 days
            per_day = emp.salary_amount / Decimal("6.00")
            calculated_amount = per_day * Decimal(present_days)
            notes = f"Auto-calculated: worked {present_days}/{total_working_days} days. Weekly Base: ₹{emp.salary_amount:,.2f}"

        calculated_amount = calculated_amount.quantize(Decimal("0.01"))

        # Check if record already exists
        record, created = SalaryRecord.objects.get_or_create(
            employee=emp,
            month=month,
            year=year,
            defaults={
                "amount": calculated_amount,
                "type": emp.salary_type,
                "status": "pending",
                "notes": notes
            }
        )

        if not created:
            record.amount = calculated_amount
            record.notes = notes
            record.type = emp.salary_type
            record.save()
            updated_count += 1
        else:
            generated_count += 1

    return Response({
        "message": f"Successfully calculated payroll for {month}/{year}.",
        "generated": generated_count,
        "updated": updated_count
    })



# --- Leave ---


def _leave_with_name(record: LeaveRequest) -> dict:
    emp = getattr(record, "employee", None)
    name = f"{emp.first_name} {emp.last_name}" if emp else _employee_name(record.employee_id)
    return leave_request_json(record, name)


@api_view(["GET", "POST"])
def leave_requests(request: Request) -> Response:
    if request.method == "GET":
        return require_auth(_leave_requests_list)(request)
    return require_auth(_leave_requests_create)(request)


def _resolve_employee_filter(params) -> int | None:
    """Return employee pk from either ?employeeId=N or ?employeeCode=XXXX."""
    if code := params.get("employeeCode") or params.get("employee_code"):
        emp = Employee.objects.filter(employee_code=code).first()
        return emp.id if emp else None
    if eid := params.get("employeeId") or params.get("employee_id"):
        return int(eid)
    return None


def _leave_requests_list(request: Request) -> Response:
    qs = LeaveRequest.objects.select_related("employee__department", "employee__designation").order_by("-id")
    qs = scope_to_branch(qs, request, field="employee__branch_id")
    employee_id = _resolve_employee_filter(request.query_params)
    leave_status = request.query_params.get("status")
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    if leave_status:
        qs = qs.filter(status=leave_status)
    return Response([_leave_with_name(r) for r in qs])


def _count_leave_days(start_str, end_str) -> int:
    """Count working days (Mon–Sat) between start and end inclusive."""
    try:
        start = date.fromisoformat(str(start_str))
        end   = date.fromisoformat(str(end_str))
    except Exception:
        return 1
    count = 0
    cur = start
    while cur <= end:
        if cur.weekday() != 6:   # skip Sunday
            count += 1
        cur += timedelta(days=1)
    return max(1, count)


def _leave_requests_create(request: Request) -> Response:
    data = request.data
    # Accept employeeCode, camelCase, or snake_case
    employee_id = None
    if code := data.get("employeeCode") or data.get("employee_code"):
        emp = Employee.objects.filter(employee_code=code).first()
        employee_id = emp.id if emp else None
    if not employee_id:
        employee_id = data.get("employeeId") or data.get("employee_id")
    start_date  = data.get("startDate")  or data.get("start_date")
    end_date    = data.get("endDate")    or data.get("end_date") or start_date
    leave_type  = data.get("type") or data.get("leave_type", "casual")

    if not employee_id or not start_date:
        return Response({"error": "employeeId and startDate are required"}, status=400)

    total_days = _count_leave_days(start_date, end_date)
    record = LeaveRequest.objects.create(
        employee_id=employee_id,
        type=leave_type,
        start_date=start_date,
        end_date=end_date,
        total_days=total_days,
        reason=data.get("reason"),
    )
    return Response(_leave_with_name(record), status=201)


@api_view(["PATCH"])
@require_hr
def update_leave_status(request: Request, pk: int) -> Response:
    from .models import LeaveBalance
    record = LeaveRequest.objects.filter(pk=pk).first()
    if not record:
        return _error("Not found", 404)

    old_status = record.status
    new_status = request.data.get("status", old_status)
    record.status = new_status
    if "hrComment" in request.data:
        record.hr_comment = request.data["hrComment"]
    record.save()

    if new_status == "approved" and old_status != "approved":
        days = Decimal(str(record.total_days or 1))
        try:
            year = int(record.start_date[:4])
        except Exception:
            year = date.today().year
        qs = LeaveBalance.objects.filter(employee_id=record.employee_id, year=year)
        if record.leave_type_ref_id:
            qs = qs.filter(leave_type_id=record.leave_type_ref_id)
        for lb in qs:
            lb.used = Decimal(str(lb.used)) + days
            lb.remaining = max(Decimal("0"), Decimal(str(lb.remaining)) - days)
            lb.save()
        Notification.objects.create(
            employee_id=record.employee_id,
            type="leave",
            message=f"Your leave request ({record.start_date} → {record.end_date}) has been approved.",
        )

    elif new_status == "rejected" and old_status != "rejected":
        Notification.objects.create(
            employee_id=record.employee_id,
            type="leave",
            message=f"Your leave request ({record.start_date} → {record.end_date}) has been rejected.",
        )

    return Response(_leave_with_name(record))


@api_view(["DELETE"])
@require_hr
def delete_leave_request(request: Request, pk: int) -> Response:
    record = LeaveRequest.objects.filter(pk=pk).first()
    if not record:
        return _error("Not found", 404)
    record.delete()
    return Response(status=204)


# --- Notifications ---


def _notif_with_name(record: Notification) -> dict:
    return notification_json(record, _employee_name(record.employee_id))


@api_view(["GET", "POST"])
@require_auth
def notifications(request: Request) -> Response:
    if request.method == "GET":
        qs = Notification.objects.select_related("employee")
        # An employee token only ever sees their own notifications — HR sees
        # everything (used for admin/debug, not a normal HR-portal screen).
        employee_id = get_token_employee_id(request)
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        if request.query_params.get("unreadOnly") in ("true", "1", True):
            qs = qs.filter(is_read=False)
        rows = [_notif_with_name(r) for r in qs]
        rows.sort(key=lambda r: r["createdAt"] or "", reverse=True)
        return Response(rows)
    return _notifications_create(request)


def _notifications_create(request: Request) -> Response:
    data = request.data
    record = Notification.objects.create(
        employee_id=data.get("employeeId"),
        type=data.get("type", "general"),
        message=data.get("message"),
    )
    return Response(_notif_with_name(record), status=201)


@api_view(["PATCH"])
@require_auth
def mark_notification_read(_request: Request, pk: int) -> Response:
    record = Notification.objects.filter(pk=pk).first()
    if not record:
        return _error("Not found", 404)
    record.is_read = True
    record.save(update_fields=["is_read"])
    return Response(_notif_with_name(record))


@api_view(["POST"])
@require_auth
def register_push_token(request: Request) -> Response:
    """
    Mobile app only — the web app has no push equivalent. Called once after
    login (and again if Expo issues a new token). Upserts by token value so
    re-registering the same device just reassigns it, which also covers
    "a different employee logged into this phone".
    """
    employee_id = get_token_employee_id(request)
    if not employee_id:
        return _error("Employee access required", 403)

    token = (request.data.get("token") or "").strip()
    if not token:
        return _error("token is required")

    PushToken.objects.update_or_create(
        token=token, defaults={"employee_id": employee_id, "platform": request.data.get("platform", "expo")}
    )
    return Response({"message": "Push token registered"}, status=201)


# --- Jobs ---


def _job_with_meta(job: Job) -> dict:
    dept_name = job.department.name if job.department_id else None
    applicant_count = Applicant.objects.filter(job_id=job.id).count()
    return job_json(job, dept_name, applicant_count)


@api_view(["GET", "POST"])
def jobs(request: Request) -> Response:
    if request.method == "GET":
        jobs_qs = Job.objects.select_related("department").order_by("id")
        return Response([_job_with_meta(j) for j in jobs_qs])
    return require_hr(_jobs_create)(request)


def _jobs_create(request: Request) -> Response:
    data = request.data
    job = Job.objects.create(
        title=data.get("title"),
        department_id=data.get("departmentId"),
        description=data.get("description"),
        requirements=data.get("requirements"),
        salary_range=data.get("salaryRange"),
        status=data.get("status", "open"),
    )
    job = Job.objects.select_related("department").get(pk=job.pk)
    return Response(_job_with_meta(job), status=201)


@api_view(["GET", "PATCH", "DELETE"])
def job_detail(request: Request, pk: int) -> Response:
    if request.method == "GET":
        job = Job.objects.select_related("department").filter(pk=pk).first()
        if not job:
            return _error("Job not found", 404)
        return Response(_job_with_meta(job))
    if request.method == "DELETE":
        return require_hr(_jobs_delete)(request, pk)
    return require_hr(_jobs_update)(request, pk)


def _jobs_update(request: Request, pk: int) -> Response:
    job = Job.objects.select_related("department").filter(pk=pk).first()
    if not job:
        return _error("Not found", 404)
    for key, attr in [
        ("title", "title"),
        ("departmentId", "department_id"),
        ("description", "description"),
        ("requirements", "requirements"),
        ("salaryRange", "salary_range"),
        ("status", "status"),
    ]:
        if key in request.data:
            setattr(job, attr, request.data[key])
    job.save()
    return Response(_job_with_meta(job))


def _jobs_delete(_request: Request, pk: int) -> Response:
    Job.objects.filter(id=pk).delete()
    return Response({"message": "Job deleted"})


# --- Applicants ---


def _applicant_with_title(applicant: Applicant) -> dict:
    job = Job.objects.filter(id=applicant.job_id).first()
    return applicant_json(applicant, job.title if job else None)


@api_view(["GET", "POST"])
def applicants(request: Request) -> Response:
    if request.method == "GET":
        return require_hr(_applicants_list)(request)
    return _applicants_submit(request)


def _applicants_list(request: Request) -> Response:
    qs = Applicant.objects.select_related("job").order_by("-id")
    job_id = request.query_params.get("jobId")
    applicant_status = request.query_params.get("status")
    if job_id:
        qs = qs.filter(job_id=int(job_id))
    if applicant_status:
        qs = qs.filter(status=applicant_status)
    return Response([_applicant_with_title(a) for a in qs])


def _applicants_submit(request: Request) -> Response:
    data = request.data
    job_id = data.get("jobId")
    if not job_id:
        return _error("jobId is required", 400)
    job = Job.objects.filter(pk=job_id).first()
    if not job:
        return _error("Job not found", 404)
    if job.status != "open":
        return _error("This position is no longer accepting applications", 400)
    if not data.get("name") or not data.get("email") or not data.get("phone"):
        return _error("Name, email, and phone are required", 400)
    applicant = Applicant.objects.create(
        job_id=job_id,
        name=data.get("name"),
        email=data.get("email"),
        phone=data.get("phone"),
        cover_letter=data.get("coverLetter"),
        experience=data.get("experience"),
    )
    return Response(_applicant_with_title(applicant), status=201)


@api_view(["PATCH"])
@require_hr
def update_applicant_status(request: Request, pk: int) -> Response:
    applicant = Applicant.objects.filter(pk=pk).first()
    if not applicant:
        return _error("Not found", 404)
    applicant.status = request.data.get("status", applicant.status)
    if "interviewDate" in request.data:
        applicant.interview_date = request.data["interviewDate"]
    if "notes" in request.data:
        applicant.notes = request.data["notes"]
    applicant.save()
    return Response(_applicant_with_title(applicant))


# --- Attendance ---


@api_view(["GET", "POST"])
def attendance(request: Request) -> Response:
    if request.method == "GET":
        return require_auth(_attendance_list)(request)
    return require_hr(_attendance_create)(request)


def _attendance_list(request: Request) -> Response:
    qs = Attendance.objects.order_by("-id")
    employee_id = request.query_params.get("employeeId")
    year = request.query_params.get("year")
    if employee_id:
        qs = qs.filter(employee_id=int(employee_id))
    if year:
        qs = qs.filter(date__startswith=str(year))
    return Response([attendance_json(r) for r in qs])


def _attendance_create(request: Request) -> Response:
    data = request.data
    record, _created = Attendance.objects.update_or_create(
        employee_id=data.get("employeeId"),
        date=data.get("date"),
        defaults={
            "present": data.get("present", True),
            "hours_worked": parse_decimal(data.get("hoursWorked")),
            "notes": data.get("notes"),
        },
    )
    return Response(attendance_json(record), status=201)


# --- Dashboard ---


MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


@api_view(["GET"])
@require_hr
def hr_dashboard_summary(_request: Request) -> Response:
    emp_stats = Employee.objects.aggregate(
        total=Count("id"),
        active=Count("id", filter=Q(status="active")),
        inactive=Count("id", filter=Q(status="inactive")),
    )
    pending_leaves = LeaveRequest.objects.filter(status="pending").count()
    unread_notifications = Notification.objects.filter(is_read=False).count()
    total_departments = Department.objects.count()
    _zero = Value(Decimal("0"), output_field=DecimalField())
    salary_stats = Employee.objects.filter(status="active").aggregate(
        monthly_total=Coalesce(
            Sum("salary_amount", filter=Q(salary_type="monthly")), _zero
        ),
        weekly_total=Coalesce(
            Sum("salary_amount", filter=Q(salary_type="weekly")), _zero
        ),
    )
    open_jobs = Job.objects.filter(status="open").count()
    pending_applicants = Applicant.objects.filter(status="applied").count()
    gender_stats = Employee.objects.filter(status="active").aggregate(
        male=Count("id", filter=Q(gender="male")),
        female=Count("id", filter=Q(gender="female")),
        other=Count("id", filter=~Q(gender__in=["male", "female"])),
    )

    today = date.today()
    geo_punches_today = AttendanceLog.objects.filter(date=today, source="geo:auto").count()
    on_duty_pending = OnDutySession.objects.filter(
        status__in=[OnDutySession.STATUS_PENDING_HOD, OnDutySession.STATUS_PENDING_HR]
    ).count()
    on_duty_sessions_active = OnDutySession.objects.filter(status=OnDutySession.STATUS_ACTIVE).count()
    on_duty_completed_today = OnDutySession.objects.filter(
        status=OnDutySession.STATUS_COMPLETED, completed_at__date=today
    ).count()
    employees_on_duty_today = OnDutySession.objects.filter(
        Q(created_at__date=today) | Q(status=OnDutySession.STATUS_ACTIVE)
    ).values("employee_id").distinct().count()
    pending_punch_verifications = OnDutyPunchVerification.objects.filter(
        status=OnDutyPunchVerification.STATUS_PENDING
    ).count()
    live_tracking_enabled = Employee.objects.filter(location_tracking_enabled=True, status="active").count()

    return Response(
        {
            "totalEmployees": emp_stats["total"] or 0,
            "activeEmployees": emp_stats["active"] or 0,
            "inactiveEmployees": emp_stats["inactive"] or 0,
            "pendingLeaves": pending_leaves,
            "unreadNotifications": unread_notifications,
            "totalDepartments": total_departments,
            "monthlySalaryTotal": float(salary_stats["monthly_total"] or 0),
            "weeklySalaryTotal": float(salary_stats["weekly_total"] or 0),
            "openJobs": open_jobs,
            "pendingApplicants": pending_applicants,
            "maleEmployees": gender_stats["male"] or 0,
            "femaleEmployees": gender_stats["female"] or 0,
            "otherEmployees": gender_stats["other"] or 0,
            "geoPunchesToday": geo_punches_today,
            "onDutyPendingApprovals": on_duty_pending,
            "onDutySessionsActive": on_duty_sessions_active,
            "onDutyCompletedToday": on_duty_completed_today,
            "employeesOnDutyToday": employees_on_duty_today,
            "pendingPunchVerifications": pending_punch_verifications,
            "liveTrackingEnabledCount": live_tracking_enabled,
        }
    )


@api_view(["GET"])
@require_auth
def employee_dashboard_summary(request: Request) -> Response:
    employee_id = request.query_params.get("employeeId") or request.jwt_user.get("employeeId")
    if not employee_id:
        return _error("employeeId required", 400)
    employee_id = int(employee_id)

    today = date.today()
    month, year = today.month, today.year
    prefix = f"{year}-{str(month).zfill(2)}"

    # Present days — from biometric logs + manual attendance
    present_dates: set = set()
    for d in AttendanceLog.objects.filter(
        employee_id=employee_id, date__year=year, date__month=month,
    ).values_list("date", flat=True).distinct():
        present_dates.add(d)
    for d_str in Attendance.objects.filter(
        employee_id=employee_id, date__startswith=prefix, present=True,
    ).values_list("date", flat=True):
        try:
            present_dates.add(date.fromisoformat(d_str))
        except Exception:
            pass
    present_days = len(present_dates)

    # Working days so far this month (Mon–Sat, up to today)
    working_days_so_far = sum(
        1 for d in range(1, today.day + 1)
        if date(year, month, d).weekday() != 6
    )

    # Approved leave days this month
    leave_days_this_month = 0
    approved_leaves = LeaveRequest.objects.filter(employee_id=employee_id, status="approved")
    for lv in approved_leaves:
        try:
            start = date.fromisoformat(str(lv.start_date))
            end   = date.fromisoformat(str(lv.end_date))
            cur   = start
            while cur <= end:
                if cur.year == year and cur.month == month and cur <= today:
                    leave_days_this_month += 1
                cur += timedelta(days=1)
        except Exception:
            pass

    absent_days = max(0, working_days_so_far - present_days - leave_days_this_month)

    # Leave balance (sum of remaining across all leave types this year)
    leave_balance = LeaveBalance.objects.filter(
        employee_id=employee_id, year=year,
    ).aggregate(total=Sum("remaining"))["total"] or 0

    # Pending requests = pending leaves + pending permissions
    pending_leaves = LeaveRequest.objects.filter(employee_id=employee_id, status="pending").count()
    pending_perms  = EmployeePermission.objects.filter(employee_id=employee_id, status="pending").count()

    payrolls = Payroll.objects.filter(employee_id=employee_id).order_by("-year", "-month")[:6]
    if payrolls.exists():
        recent = [_salary_from_payroll(p) for p in payrolls]
    else:
        recent = [
            _salary_with_name(r)
            for r in SalaryRecord.objects.filter(employee_id=employee_id)
            .order_by("-year", "-month")[:6]
        ]

    # Manager access flags
    from .models import DepartmentManager
    from django.db.models import Q as DQ
    manager_profile = None
    try:
        manager_profile = DepartmentManager.objects.prefetch_related(
            "department_assignments", "employee_assignments"
        ).get(employee_id=employee_id, is_active=True)
    except DepartmentManager.DoesNotExist:
        pass

    is_manager = manager_profile is not None
    pending_approvals_count = 0
    if is_manager:
        dept_ids = [da.department_id for da in manager_profile.department_assignments.all()]
        direct_ids = [ea.employee_id for ea in manager_profile.employee_assignments.all()]
        emp_filter = DQ(employee_id__in=direct_ids)
        if dept_ids:
            emp_filter |= DQ(employee__department_id__in=dept_ids)
        pending_approvals_count = (
            LeaveRequest.objects.filter(emp_filter, status="pending").count()
            + EmployeePermission.objects.filter(emp_filter, status="pending").count()
        )

    return Response({
        "employeeId":     employee_id,
        "presentDays":    present_days,
        "absentDays":     absent_days,
        "leaveDays":      leave_days_this_month,
        "leaveBalance":   float(leave_balance),
        "pendingRequests": pending_leaves + pending_perms,
        "pendingLeaves":  pending_leaves,
        "approvedLeaves": approved_leaves.count(),
        "recentSalaries": recent,
        "isManager":      is_manager,
        "canSubmitLeave": is_manager,
        "pendingApprovalsCount": pending_approvals_count,
    })


@api_view(["GET"])
@require_hr
def interview_summary(_request: Request) -> Response:
    stats = {
        row["status"]: row["count"]
        for row in Applicant.objects.values("status").annotate(count=Count("id"))
    }
    total = sum(stats.values())
    return Response(
        {
            "totalApplicants": total,
            "attended": stats.get("attended", 0),
            "selected": stats.get("selected", 0),
            "rejected": stats.get("rejected", 0),
            "pending": stats.get("applied", 0),
        }
    )


@api_view(["GET"])
@require_hr
def salary_trends(_request: Request) -> Response:
    # Prefer Payroll table (authoritative) when present, otherwise fall back to legacy SalaryRecord
    _zero = Value(Decimal("0"), output_field=DecimalField())

    payroll_qs = (
        Payroll.objects.values("month", "year")
        .annotate(total=Coalesce(Sum("final_salary"), _zero))
        .order_by("year", "month")[:12]
    )

    if payroll_qs.exists():
        trends = payroll_qs
    else:
        trends = (
            SalaryRecord.objects.values("month", "year")
            .annotate(total=Coalesce(Sum("amount"), _zero))
            .order_by("year", "month")[:12]
        )

    return Response(
        [
            {
                "month": t["month"],
                "year": t["year"],
                "total": float(t["total"]),
                "label": f"{MONTH_NAMES[t['month'] - 1]} {t['year']}",
            }
            for t in trends
        ]
    )

