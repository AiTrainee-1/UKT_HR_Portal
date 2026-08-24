"""
Mobile App Login -HR-facing visibility into who can actually get into the
employee mobile app, and a way to fix them when they can't.

On passwords: `Employee.password_hash` is a bcrypt hash, which is one-way by
construction -there is no operation, here or anywhere else, that turns it
back into the password the employee chose. So this module exposes whether a
password *exists* and offers to replace it; it never returns the password
itself, because it cannot. Resetting is also the right answer to the problem
this page exists to solve ("employee forgot their password"): HR sets a new
one and tells them, or clears it so they run Set Password again themselves.

Two different signals are reported per employee, and they mean different
things:
  - hasPassword          -has completed Set Password, so an account exists.
                          This is the historical "has mobile access" signal.
  - lastMobileLoginAt    -stamped by employee_login() on each sign-in. Only
                          populated from the release that added the field
                          onwards, so an employee with a password but no
                          recorded login has simply not signed in *since*
                          then, not necessarily never.
"""
from __future__ import annotations

import io

import bcrypt
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework.decorators import api_view
from rest_framework.request import Request
from rest_framework.response import Response

from .audit_utils import log_action
from .auth import require_hr
from .branch_scope import scope_to_branch
from .models import Employee, PushToken

MIN_PASSWORD_LEN = 8


def _employee_json(emp: Employee, device_counts: dict[int, int]) -> dict:
    has_password = bool(emp.password_hash)
    return {
        "id": emp.id,
        "employeeCode": emp.employee_code,
        "name": f"{emp.first_name} {emp.last_name}".strip(),
        "department": emp.department.name if emp.department else None,
        "designation": emp.designation.title if emp.designation else None,
        "phone": emp.phone or None,
        "email": emp.email or None,
        "status": emp.status,
        "employmentType": emp.employment_type,
        "hasPassword": has_password,
        "lastMobileLoginAt": (
            emp.last_mobile_login_at.isoformat() if emp.last_mobile_login_at else None
        ),
        # A registered push token means the app has actually run on a device
        # while signed in -useful corroboration that the account is in real
        # use rather than just provisioned.
        "deviceCount": device_counts.get(emp.id, 0),
    }


def _filtered(request: Request) -> tuple[list[Employee], list[Employee], list[Employee], list[Employee], list[Employee]]:
    """
    Applies this page's shared filtering and returns
    (base, has_access, no_access, signed_in, rows-for-the-active-tab).

    Scoped to STAFF only -production employees are out of scope for this page
    and are never listed or counted here, so the numbers on it will not match
    the Employees page's org-wide totals.

    Query params:
      access = all | has_access | no_access | signed_in | never_signed_in
      status = all | active | inactive
      search = matches employee code, first/last name, phone or email
    """
    access = (request.query_params.get("access") or "all").strip()
    status_filter = (request.query_params.get("status") or "all").strip()
    search = (request.query_params.get("search") or "").strip()

    qs = (
        scope_to_branch(Employee.objects, request)
        .filter(employment_type="staff")
        .select_related("department", "designation")
    )

    if status_filter in ("active", "inactive"):
        qs = qs.filter(status=status_filter)

    if search:
        qs = qs.filter(
            Q(employee_code__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(phone__icontains=search)
            | Q(email__icontains=search)
        )

    # Counts are computed over the status/search-filtered set but *before* the
    # access filter, so the summary cards stay stable while HR clicks between
    # access tabs rather than each tab reporting only itself.
    base = list(qs.order_by("employee_code"))

    has_access = [e for e in base if e.password_hash]
    no_access = [e for e in base if not e.password_hash]
    signed_in = [e for e in has_access if e.last_mobile_login_at]

    if access == "has_access":
        rows = has_access
    elif access == "no_access":
        rows = no_access
    elif access == "signed_in":
        rows = signed_in
    elif access == "never_signed_in":
        rows = [e for e in base if not e.last_mobile_login_at]
    else:
        rows = base

    return base, has_access, no_access, signed_in, rows


@api_view(["GET"])
@require_hr
def mobile_app_logins(request: Request) -> Response:
    """Staff employees with their mobile-app access state, plus headline counts."""
    base, has_access, no_access, signed_in, rows = _filtered(request)

    device_counts: dict[int, int] = {
        row["employee_id"]: row["n"]
        for row in PushToken.objects.values("employee_id").annotate(n=Count("id"))
    }

    return Response({
        "summary": {
            "total": len(base),
            "hasAccess": len(has_access),
            "noAccess": len(no_access),
            "signedIn": len(signed_in),
            "activeNoAccess": sum(1 for e in no_access if e.status == "active"),
        },
        "results": [_employee_json(e, device_counts) for e in rows],
    })


ACCESS_LABELS = {
    "no_access": "Never Set Up",
    "has_access": "Has App Access",
    "signed_in": "Signed In",
    "never_signed_in": "No Sign-In Recorded",
    "all": "All Staff",
}


@api_view(["GET"])
@require_hr
def mobile_app_logins_export(request: Request) -> HttpResponse:
    """
    The currently-filtered staff list as an .xlsx download.

    Honours exactly the same access/status/search params as the list endpoint
    (via _filtered), so what downloads is what's on screen -no second filter
    to keep in sync.
    """
    _, _, _, _, rows = _filtered(request)
    access = (request.query_params.get("access") or "all").strip()
    label = ACCESS_LABELS.get(access, "Staff")

    wb = Workbook()
    ws = wb.active
    # Excel caps sheet names at 31 chars and rejects []:*?/\
    ws.title = label[:31]

    headers = ["Name", "Employee Code", "Phone Number", "Department"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="006496")  # UKT Blue
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for emp in rows:
        ws.append([
            f"{emp.first_name} {emp.last_name}".strip(),
            # Codes are digit strings like "2580"; written as text so Excel
            # doesn't reformat them as numbers and drop any leading zeros.
            str(emp.employee_code or ""),
            str(emp.phone or ""),
            emp.department.name if emp.department else "",
        ])

    for col, width in zip("ABCD", (30, 16, 18, 26)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"mobile-app-login-{access}-{timezone.localdate().isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_action(
        request, "export", "mobile_app_login",
        description=f"Exported {len(rows)} staff ({label}) from Mobile App Login",
    )
    return response


@api_view(["POST"])
@require_hr
def mobile_app_reset_password(request: Request, employee_id: int) -> Response:
    """
    Replace (or clear) one employee's mobile-app password.

    Body: {"password": "<new password>"}  -> sets it, employee can log in now
          {"clear": true}                 -> wipes it, employee must run
                                             Set Password in the app again

    Never returns the stored password (see module docstring -it's a bcrypt
    hash and cannot be recovered). The new password is echoed back only when
    HR supplied it, since HR already knows it and has to pass it on.
    """
    emp = scope_to_branch(Employee.objects, request).filter(pk=employee_id).first()
    if not emp:
        return Response({"error": "Employee not found"}, status=404)

    if request.data.get("clear"):
        if not emp.password_hash:
            return Response({"error": "This employee has no password set."}, status=400)
        emp.password_hash = None
        emp.save(update_fields=["password_hash", "updated_at"])
        log_action(
            request, "update", "mobile_app_login", record_id=emp.id,
            description=f"Cleared mobile app password for {emp.employee_code} ({emp.first_name} {emp.last_name})",
        )
        return Response({
            "message": "Password cleared. The employee must set a new password from the app.",
            "hasPassword": False,
        })

    password = request.data.get("password") or ""
    if len(password) < MIN_PASSWORD_LEN:
        return Response(
            {"error": f"Password must be at least {MIN_PASSWORD_LEN} characters."}, status=400
        )

    was_set = bool(emp.password_hash)
    emp.password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt(rounds=10)).decode()
    # A password reset is an access change, not a sign-in -deliberately does
    # not touch last_mobile_login_at, so "has HR fixed them yet vs have they
    # actually got back in" stays answerable.
    emp.save(update_fields=["password_hash", "updated_at"])

    log_action(
        request, "update", "mobile_app_login", record_id=emp.id,
        description=(
            f"{'Reset' if was_set else 'Set'} mobile app password for "
            f"{emp.employee_code} ({emp.first_name} {emp.last_name})"
        ),
    )
    return Response({
        "message": f"Password {'reset' if was_set else 'set'} for {emp.first_name} {emp.last_name}.",
        "hasPassword": True,
        "employeeId": emp.id,
    })
